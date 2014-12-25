# coding: utf-8
from openpyxl import load_workbook, Workbook

__author__ = 'valiullin'


COLUMN = 1
VALUE = 'XXX'
xl_file = 's.xlsx'
out_file = "x_out.xlsx"

wb = load_workbook(filename=xl_file, use_iterators=True)
ws = wb.get_sheet_by_name(name='Sheet1')

wb_dest = Workbook(optimized_write=True)
ws_dest = wb_dest.create_sheet()

for row in ws.iter_rows():
    if row[COLUMN].internal_value == VALUE:
        raw_row = [v.internal_value for v in row]
        ws_dest.append(raw_row)

# Save the file
wb_dest.save(out_file)